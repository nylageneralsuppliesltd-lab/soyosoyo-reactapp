from datetime import date, timedelta
import csv
import os

start = date(2026, 2, 24)
rows = []
headers = [
    'Phase','Week','Workstream','Module','Task','Priority','EffortHours','OwnerRole',
    'Dependencies','DefinitionOfDone','TargetStart','TargetEnd','Status','Milestone','CommercialImpact'
]

def add(phase, week, ws, module, task, prio, hrs, owner, deps, dod, s, e, status, ms, impact):
    rows.append([phase, week, ws, module, task, prio, hrs, owner, deps, dod, s.isoformat(), e.isoformat(), status, ms, impact])

w1s = start
w1e = start + timedelta(days=6)
w2s = w1e + timedelta(days=1)
w2e = w2s + timedelta(days=6)
w3s = w2e + timedelta(days=1)
w3e = w3s + timedelta(days=6)
w4s = w3e + timedelta(days=1)
w4e = w4s + timedelta(days=6)
w5s = w4e + timedelta(days=1)
w5e = w5s + timedelta(days=6)
w6s = w5e + timedelta(days=1)
w6e = w6s + timedelta(days=6)
w7s = w6e + timedelta(days=1)
w7e = w7s + timedelta(days=6)
w8s = w7e + timedelta(days=1)
w8e = w8s + timedelta(days=6)
w9s = w8e + timedelta(days=1)
w9e = w9s + timedelta(days=6)
w10s = w9e + timedelta(days=1)
w10e = w10s + timedelta(days=6)

add('Phase 0',1,'Platform Stability','Backend Runtime','Fix backend start:dev and standardize env/start scripts','P0',10,'Fullstack Dev','Current backend config','Backend runs on port 3000 reliably for 3 consecutive restarts',w1s,w1e,'Planned','M0 Runtime stable','Required for demos and onboarding')
add('Phase 0',1,'Platform Stability','Observability','Add startup health checks and error logging baseline','P0',4,'Backend Dev','Backend runtime stable','/health and structured startup errors available',w1s,w1e,'Planned','M0 Runtime stable','Reduces support risk during pilots')
add('Phase 0',1,'Core Finance QA','Deposits/Withdrawals','Complete smoke suite for all payment types and reversals','P0',8,'QA + Fullstack','Latest deposits fixes','20/20 high-value tests pass',w1s,w1e,'Planned','M1 Cashflow hardened','Protects accounting trust')

add('Phase 1',2,'Security & Access','Users & Roles','Implement user profiles and role CRUD','P0',12,'Fullstack Dev','Runtime stable','Admin can create/edit/disable users and assign roles',w2s,w2e,'Planned','M2 Access foundation','Needed for multi-SACCO sale readiness')
add('Phase 1',2,'Security & Access','Permissions','Implement permission matrix per module/action','P0',12,'Backend Dev','Users & Roles','Unauthorized actions blocked server-side and hidden client-side',w2s,w2e,'Planned','M2 Access foundation','Enterprise requirement')
add('Phase 1',3,'Approvals','Maker-Checker','Add configurable approvals for high-risk actions','P0',14,'Fullstack Dev','Permissions','Pending/approved/rejected workflow with audit trail',w3s,w3e,'Planned','M3 Governance controls','Compliance + controls')
add('Phase 1',3,'Audit','Audit Trail','Immutable audit entries for settings and financial actions','P0',8,'Backend Dev','Maker-Checker','All critical actions include actor, before/after, timestamp',w3s,w3e,'Planned','M3 Governance controls','Critical in due diligence')

add('Phase 2',4,'Assets Lifecycle','Asset Purchases','Create asset purchase posting flow with capitalization','P1',12,'Fullstack Dev','Permissions + audit','Purchase posts correctly to asset and cash/bank accounts',w4s,w4e,'Planned','M4 Asset purchases live','Expands product scope')
add('Phase 2',4,'Assets Lifecycle','Depreciation','Automate depreciation schedules and monthly posting','P1',10,'Backend Dev','Asset purchases','Depreciation run generates expected journal entries',w4s,w4e,'Planned','M4 Asset purchases live','Improves reporting quality')
add('Phase 2',5,'Assets Lifecycle','Asset Disposal/Sale','Implement sale/disposal with gain/loss accounting','P1',12,'Fullstack Dev','Asset purchases + depreciation','Disposal updates register and posts gain/loss',w5s,w5e,'Planned','M5 Asset lifecycle complete','Strong differentiator for SACCOs')
add('Phase 2',5,'Configuration','Settings Governance','Settings versioning + change approvals + effective dates','P1',8,'Backend Dev','Audit trail','Settings changes tracked and reversible',w5s,w5e,'Planned','M5 Asset lifecycle complete','Safer operations')

add('Phase 3',6,'Multi-SACCO','Tenant Isolation','Harden data isolation and tenancy boundaries','P0',14,'Backend Dev','User access model','No cross-tenant reads/writes in critical endpoints',w6s,w6e,'Planned','M6 Multi-SACCO hardening','Mandatory for commercialization')
add('Phase 3',6,'Multi-SACCO','Tenant Provisioning','Onboard new SACCO wizard + default charts/settings','P1',10,'Fullstack Dev','Tenant isolation','New SACCO can be provisioned in under 15 minutes',w6s,w6e,'Planned','M6 Multi-SACCO hardening','Accelerates sales onboarding')
add('Phase 3',7,'Branding','White-label','Per-SACCO branding (logo/colors/legal docs)','P2',8,'Frontend Dev','Tenant model','Branding applies cleanly by tenant',w7s,w7e,'Planned','M7 Sellable product pack','Supports resale strategy')
add('Phase 3',7,'Commercial','Packaging','Define pricing tiers + feature flags per plan','P1',8,'Product + Dev','Tenant provisioning','Feature availability configurable by plan',w7s,w7e,'Planned','M7 Sellable product pack','Revenue readiness')

add('Phase 4',8,'Reporting','Financial Reports','Finalize BS/IS/TB reconciliation packs and exports','P0',10,'Finance QA + Dev','Core modules stable','Month-end pack ties out with zero unexplained variance',w8s,w8e,'Planned','M8 Reporting confidence','Client trust and auditability')
add('Phase 4',8,'Reporting','Management Dashboards','Operational KPI dashboards for SACCO leadership','P2',8,'Frontend Dev','Reporting data quality','KPI cards and trends validated',w8s,w8e,'Planned','M8 Reporting confidence','Executive visibility')
add('Phase 4',9,'Integrations','Bank/MPESA Imports','Import and reconcile bank + mobile money statements','P1',12,'Fullstack Dev','Tenant + permissions','Import parser + reconciliation workflow complete',w9s,w9e,'Planned','M9 Integration layer','Reduces manual work')
add('Phase 4',9,'Integrations','Notifications','Email/SMS notifications for approvals and arrears','P2',8,'Backend Dev','Approvals','Templates + send logs + retries',w9s,w9e,'Planned','M9 Integration layer','Improves collections')

add('Phase 5',10,'Quality','UAT + Regression','Run UAT scripts across all core modules','P0',12,'QA + Product','All prior milestones','UAT pass rate >= 95% with no open P0',w10s,w10e,'Planned','M10 Release candidate','Go-live readiness')
add('Phase 5',10,'Operations','Go-live Playbook','Deploy, backup/restore drill, incident playbook','P0',8,'DevOps + Lead','UAT complete','Runbook approved and restore tested',w10s,w10e,'Planned','M10 Release candidate','Operational confidence')
add('Phase 5',10,'Commercial','Sales Enablement','Create demo script, one-pager, implementation checklist','P1',6,'Product + Lead','Stable release candidate','Reusable sales/demo assets ready',w10s,w10e,'Planned','M10 Release candidate','Supports client acquisition')

base = os.path.dirname(os.path.dirname(__file__))
csv_path = os.path.join(base, 'PROJECT_360_MILESTONE_SCHEDULE.csv')
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(headers)
    w.writerows(rows)

xlsx_path = os.path.join(base, 'PROJECT_360_MILESTONE_SCHEDULE.xlsx')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Milestones'

    ws.append(headers)
    for r in rows:
        ws.append(r)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    widths = [12,7,18,18,52,8,11,14,22,42,12,12,10,22,28]
    for i, wid in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wid

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.freeze_panes = 'A2'

    summary = wb.create_sheet('Summary')
    summary.append(['Metric', 'Value'])
    summary.append(['Schedule Duration (weeks)', '10'])
    summary.append(['Part-time Capacity (hrs/week)', '12-15'])
    summary.append(['Total Planned Effort (hrs)', str(sum(int(r[6]) for r in rows))])
    summary.append(['P0 Items', str(sum(1 for r in rows if r[5] == 'P0'))])
    summary.append(['P1 Items', str(sum(1 for r in rows if r[5] == 'P1'))])
    summary.append(['P2 Items', str(sum(1 for r in rows if r[5] == 'P2'))])
    summary.append(['Commercial Milestone', 'Multi-SACCO hardened + sales pack'])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    summary.column_dimensions['A'].width = 35
    summary.column_dimensions['B'].width = 50

    wb.save(xlsx_path)
    print('Created:', csv_path)
    print('Created:', xlsx_path)
except Exception as exc:
    print('Created:', csv_path)
    print('XLSX not created:', exc)
